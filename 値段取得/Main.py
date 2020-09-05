from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
from selenium.common.exceptions import ElementClickInterceptedException, ElementNotInteractableException
from parts.login import Login
from parts.logout import Logout
import pandas as pd
import sys
import os
import datetime
import time
import openpyxl as op


now = datetime.datetime.now()
print("開始：{}".format(now))
url = "https://www.amazon.co.jp/login"
DRIVER_PATH = r"WebDriver\chromedriver.exe"
# DRIVER_PATH = r"WebDriver\chromedriver.exe"

options = Options()
options.add_argument('--disable-gpu')
options.add_argument('--disable-extensions')
options.add_argument('--proxy-server="direct://"')
options.add_argument('--proxy-bypass-list=*')
options.add_argument('--start-maximized')
Email = "*****@***.**.**"
password = "******"
sss_time_cnt = 0.5
ss_time_cnt = 1
s_time_cnt = 3
mm_time_cnt = 5
m_time_cnt = 7
l_time_cnt = 10
x_time_cnt = 15
xl_time_cnt = 20
xll_time_cnt = 30
OUTPUTpath = r"OUTPUT\yyyymm\dd"
Logfaile = "Error.log"
LogScreenshot = r"screenshot\Error_hhmmss.jpg"
read_excel = r"INPUT\商品一覧.xlsx"
quit_browser = None
header = ["商品タイトル", "金額", "レビュー点数", "リンク"]

write_excel = "Amazon_商品名_hhmmss.xlsx"


OUTPUTpath = OUTPUTpath.replace("yyyymm", now.strftime("%Y%m")).replace("dd", now.strftime("%d"))


try:

    def error_browser(message, time, error_logpath, error_logfile, screenshotpath, driver):
        if os.path.exists(error_logpath) is False:
            os.makedirs(error_logpath)

        if os.path.exists(os.path.join(error_logpath, "screenshot")) is False:
            os.makedirs(os.path.join(error_logpath, "screenshot"))

        with open(os.path.join(error_logpath, error_logfile), "a", encoding="utf_8") as f:
            f.write("・ErrorTime:" + message + "\r\n")
        screenshotpath = screenshotpath.replace("hhmmss", time)
        driver.save_screenshot(os.path.join(error_logpath, screenshotpath))

        driver.quit()
        sys.exit()

    def error_write(message, error_logpath, error_logfile):
        if os.path.exists(error_logpath) is False:
            os.makedirs(error_logpath)

        with open(os.path.join(error_logpath, error_logfile), "a", encoding="utf_8") as f:
            f.write("・ErrorTime:" + message + "\r\n")

    if os.path.exists(OUTPUTpath) is False:
        os.makedirs(OUTPUTpath)

    # ログインパージを開く
    service = Service(DRIVER_PATH)
    service.start()
    driver = webdriver.Remote(service.service_url)
    driver.get(url)
    driver.maximize_window()

    selector = "#authportal-main-section > div:nth-child(2) > div > div.a-section > form > div > div > div > h1"

    element = WebDriverWait(driver, 30).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, selector)))

    # ログイン
    login_ins = Login(driver)
    success = login_ins.input_email(Email)
    if success is not None:
        error_browser(success[1], success[0], OUTPUTpath, Logfaile, LogScreenshot, driver)

    success = login_ins.input_pass(password)

    if success is not None:
        error_browser(success[1], success[0], OUTPUTpath, Logfaile, LogScreenshot, driver)

    # Excel読みとり
    merchandise_df = pd.read_excel(read_excel)
    for merchandise in merchandise_df["商品名"]:

        main_paste_excel = []

        # 商品検索
        selector = "#twotabsearchtextbox"
        element = driver.find_element_by_css_selector(selector)
        element.clear()
        element.send_keys(merchandise)
        element.send_keys(Keys.ENTER)

        WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located)

        xpath = "//*[@id=" + '"search"' + "]/span/div/span/h1/div/div[1]/div/div/span[3]"
        element_text = driver.find_element_by_xpath(xpath).text  # ().repalce('"', '')

        if merchandise.replace(' ', '').replace('　', '').upper() == element_text.replace('"', '').replace(' ', '').replace('　', '').upper():

            class_name = "a-disabled"
            pages = int(driver.find_elements_by_class_name(class_name)[1].text)
            original_window = driver.current_window_handle

            id_name = "s-result-sort-select"
            element = driver.find_element_by_id(id_name)
            Select(element).select_by_value("review-rank")
            element.send_keys(Keys.ENTER)
            # select_by_value_text('レビューの評価順')
            WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located)
            time.sleep(mm_time_cnt)

            print(merchandise)
            class_name_idx = 0
            """
            while True:
                s_line_clamp_name = "s-line-clamp-{}".format(class_name_idx)
                elements = driver.find_elements_by_class_name(s_line_clamp_name)

                if len(elements) == 0:
                    class_name_idx += 1

                else:
                    break
            """
            for page in range(1, pages + 1):

                print("ページ{}".format(page))
                div_idx = 2
                click_flag = True

                class_name = "s-line-clamp-4"
                elements = driver.find_elements_by_class_name(class_name)

                for emt in elements:

                    class_name = "a-text-normal"
                    emts = emt.find_elements_by_class_name(class_name)
                    product_name = emts[1].text

                    if merchandise.lower() in product_name.lower():

                        emts[0].click()

                        # 商品詳細ページ
                        WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(2))
                        for window_handle in driver.window_handles:
                            if window_handle != original_window:
                                driver.switch_to.window(window_handle)
                                break
                        WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located)
                        time.sleep(ss_time_cnt)

                        try:
                            # レビュースコア取得
                            review_score = driver.find_element_by_xpath(
                                "//*[@id=" + '"reviewsMedley"' + "]/div/div[1]/div[2]/div[1]/div/div[2]/div/span/span").text

                        except NoSuchElementException:
                            pass

                        else:
                            review_score = review_score.replace('星5つ中の', '')

                            if float(review_score) >= 4:

                                # 金額取得
                                class_name = "a-color-price"  # "priceBlockBuyingPriceString"
                                price = driver.find_element_by_class_name(class_name).text

                                # URL取得
                                link_url = driver.current_url

                                row_excel = [product_name, price, review_score,
                                             '=HYPERLINK("{0}", "Go")'.format(link_url)]
                                # row_excel = [product_name, price, review_score, link_url]

                                main_paste_excel.append(row_excel)

                        driver.close()

                        WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(1))
                        driver.switch_to.window(original_window)

                class_name = "a-selected"
                element = driver.find_element_by_class_name(class_name)
                element_text = element.text

                if int(element_text) != pages:

                    class_name = "a-last"
                    element = driver.find_element_by_class_name(class_name)
                    element.click()
                    WebDriverWait(driver, 30).until(EC.presence_of_all_elements_located)
                    time.sleep(mm_time_cnt)

            out_excel_faile = "Amazon_{0}_{1}.xlsx".format(merchandise, now.strftime("%H%M%S"))
            book = op.Workbook()
            sheet = book.active

            for c, h in enumerate(header):
                sheet.cell(row=1, column=c+1).value = h

            for r, pe in enumerate(main_paste_excel):
                for c, e in enumerate(pe):
                    sheet.cell(row=r + 2, column=c + 1).value = e

            book.save(os.path.join(OUTPUTpath, out_excel_faile))
            book.close()

        else:
            print(False)

    # ログアウト

    logout_ins = Logout(driver)
    success = logout_ins.logout()
    if success is not None:
        error_browser(success[1], success[0], OUTPUTpath, Logfaile, LogScreenshot, driver)


except KeyError:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_write(error_massage, OUTPUTpath, Logfaile)

except TimeoutException:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_browser(error_massage, now.strftime("%H%M%S"),
                  OUTPUTpath, Logfaile, LogScreenshot, driver)

except WebDriverException:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_write(error_massage, OUTPUTpath, Logfaile)

except FileNotFoundError:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_write(error_massage, OUTPUTpath, Logfaile)

except NoSuchElementException:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_browser(error_massage, now.strftime("%H%M%S"),
                  OUTPUTpath, Logfaile, LogScreenshot, driver)

except ElementNotInteractableException:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_browser(error_massage, now.strftime("%H%M%S"),
                  OUTPUTpath, Logfaile, LogScreenshot, driver)

except ElementClickInterceptedException:
    import traceback
    now = datetime.datetime.now()
    error_massage = now.strftime("%H:%M:%S") + "\r\n" + traceback.format_exc().strip()
    error_browser(error_massage, now.strftime("%H%M%S"),
                  OUTPUTpath, Logfaile, LogScreenshot, driver)

finally:
    if quit_browser is None:
        driver.quit()
    now = datetime.datetime.now()
    print("終了：{}".format("now"))
